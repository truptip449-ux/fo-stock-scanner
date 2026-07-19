#!/usr/bin/env python3
"""
Market Data Manager
Handles FUTSTK filtering, Fyers API integration, and caching.
"""

import logging
import httpx
import pandas as pd
from typing import List, Dict, Optional
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import asyncio

logger = logging.getLogger(__name__)


class MarketDataManager:
    """
    Manages F&O market data: fetching, filtering, and caching.
    """
    
    NSE_FO_CSV_URL = "https://public.fyers.in/sym_details/NSE_FO.csv"
    FYERS_QUOTE_API = "https://api-t1.fyers.in/api/v3/quotes/"
    
    def __init__(self, fyers_auth):
        """
        Initialize market data manager.
        
        Args:
            fyers_auth: FyersAuthManager instance
        """
        self.fyers_auth = fyers_auth
        self.cached_data: List[Dict] = []
        self.cache_timestamp: Optional[datetime] = None
        self.master_data: Optional[pd.DataFrame] = None
        
        logger.info("MarketDataManager initialized")
    
    async def refresh_futstk_data(self) -> List[Dict]:
        """
        Main refresh pipeline:
        1. Download and parse NSE_FO master CSV
        2. Filter FUTSTK symbols only
        3. Identify 3 nearest active expiry months
        4. Fetch spot + futures quotes via Fyers API
        5. Cache and return data
        
        Returns:
            List of FUTSTK records with spot and 3 futures prices
        """
        try:
            logger.info("Starting FUTSTK refresh pipeline")
            
            # Step 1: Download master CSV
            logger.info("Downloading NSE_FO master CSV...")
            master_df = await self._download_master_csv()
            if master_df is None or len(master_df) == 0:
                raise Exception("Failed to download or parse master CSV")
            
            logger.info(f"Master CSV loaded: {len(master_df)} symbols")
            self.master_data = master_df
            
            # Step 2: Filter FUTSTK only (NO indices, NO options)
            logger.info("Filtering FUTSTK symbols...")
            futstk_df = self._filter_futstk(master_df)
            logger.info(f"Filtered FUTSTK symbols: {len(futstk_df)}")
            
            # Step 3: Get unique underlying stocks
            underlying_stocks = futstk_df['underlying'].unique().tolist()
            logger.info(f"Unique underlying stocks: {len(underlying_stocks)}")
            
            # Step 4: Build symbol mapping (spot + 3 futures for each stock)
            symbol_mapping = self._build_symbol_mapping(futstk_df, underlying_stocks)
            
            # Step 5: Fetch quotes from Fyers API
            logger.info("Fetching quotes from Fyers API...")
            scanner_data = await self._fetch_quotes_chunked(symbol_mapping)
            
            # Step 6: Cache and return
            self.cached_data = scanner_data
            self.cache_timestamp = datetime.now()
            
            logger.info(f"Refresh complete: {len(scanner_data)} records cached")
            return scanner_data
        
        except Exception as e:
            logger.error(f"FUTSTK refresh failed: {e}", exc_info=True)
            raise
    
    async def _download_master_csv(self) -> Optional[pd.DataFrame]:
        """
        Download NSE_FO master CSV from Fyers.
        
        Returns:
            DataFrame with symbol details
        """
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(self.NSE_FO_CSV_URL)
                response.raise_for_status()
                
                # Parse CSV
                csv_data = response.text
                df = pd.read_csv(io.StringIO(csv_data))
                
                logger.info(f"Master CSV downloaded: {len(df)} rows")
                return df
        
        except Exception as e:
            logger.error(f"Failed to download master CSV: {e}")
            return None
    
    def _filter_futstk(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Filter only FUTSTK symbols from master data.
        CRITICAL: NO indices (NIFTY, BANKNIFTY, etc.), NO options (CE/PE).
        
        Args:
            df: Master CSV DataFrame
            
        Returns:
            Filtered DataFrame with FUTSTK only
        """
        try:
            # Look for columns that might contain symbol type info
            # Common column names: 'ExchangeTokens', 'Symbol', 'Segment', 'ProductType'
            
            # Ensure we have the right columns
            if 'Segment' not in df.columns:
                # Infer from symbol patterns
                df['type'] = df.iloc[:, 0].apply(self._infer_symbol_type)
            else:
                df['type'] = df['Segment']
            
            # Filter: Keep only FUTSTK (Futures on stocks)
            # Exclude: FUTIDX (indices), CE/PE (options), and others
            futstk_mask = df['type'].str.contains('FUTSTK', case=False, na=False)
            futstk_df = df[futstk_mask].copy()
            
            # Extract underlying stock symbol
            if 'Symbol' in futstk_df.columns:
                futstk_df['underlying'] = futstk_df['Symbol'].str.extract(r'^([A-Z&]+)-')[0]
            elif 'ExchangeTokens' in futstk_df.columns:
                futstk_df['underlying'] = futstk_df['ExchangeTokens'].str.extract(r'^([A-Z&]+)')[0]
            else:
                # Fallback: extract from first column
                col_name = futstk_df.columns[0]
                futstk_df['underlying'] = futstk_df[col_name].str.extract(r'^([A-Z&]+)')[0]
            
            # Remove duplicates (keep unique underlyings)
            futstk_df = futstk_df.drop_duplicates(subset=['underlying'])
            
            logger.info(f"Filtered {len(futstk_df)} unique FUTSTK underlying stocks")
            return futstk_df
        
        except Exception as e:
            logger.error(f"FUTSTK filtering error: {e}")
            return df[[]]
    
    def _infer_symbol_type(self, symbol: str) -> str:
        """
        Infer symbol type from symbol string.
        
        Args:
            symbol: Symbol string
            
        Returns:
            Symbol type (FUTSTK, FUTIDX, CE, PE, etc.)
        """
        if not isinstance(symbol, str):
            return ""
        
        symbol_upper = symbol.upper()
        
        # Check for options
        if symbol_upper.endswith('CE') or symbol_upper.endswith('PE'):
            return symbol_upper[-2:]
        
        # Check for index futures (NIFTY, BANKNIFTY, etc.)
        if any(idx in symbol_upper for idx in ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY']):
            return 'FUTIDX'
        
        # Default to FUTSTK
        return 'FUTSTK'
    
    def _build_symbol_mapping(self, futstk_df: pd.DataFrame, underlying_stocks: List[str]) -> Dict[str, Dict]:
        """
        Build mapping of underlying stocks to spot + 3 futures symbols.
        
        Args:
            futstk_df: DataFrame with FUTSTK symbols
            underlying_stocks: List of unique underlying stocks
            
        Returns:
            Dict mapping underlying -> {spot_symbol, future1_symbol, future2_symbol, future3_symbol}
        """
        mapping = {}
        
        try:
            for underlying in underlying_stocks:
                try:
                    # Get all futures contracts for this underlying from master
                    futures = futstk_df[futstk_df['underlying'] == underlying]
                    
                    if len(futures) == 0:
                        continue
                    
                    # Spot symbol (NSE equity)
                    spot_symbol = f"{underlying}-EQ"
                    
                    # Get futures symbols (typically sorted by expiry)
                    futures_symbols = []
                    for _, row in futures.iterrows():
                        sym = row.iloc[0] if isinstance(row.iloc[0], str) else row.get('Symbol', '')
                        if sym and isinstance(sym, str):
                            futures_symbols.append(sym)
                    
                    # Sort by expiry (nearest first)
                    futures_symbols = sorted(futures_symbols)[:3]  # Take 3 nearest
                    
                    # Ensure we have exactly 3 futures
                    while len(futures_symbols) < 3:
                        futures_symbols.append(None)
                    
                    mapping[underlying] = {
                        'spot': spot_symbol,
                        'future1': futures_symbols[0],
                        'future2': futures_symbols[1],
                        'future3': futures_symbols[2]
                    }
                
                except Exception as e:
                    logger.warning(f"Error mapping {underlying}: {e}")
                    continue
            
            logger.info(f"Symbol mapping created: {len(mapping)} stocks")
            return mapping
        
        except Exception as e:
            logger.error(f"Symbol mapping error: {e}")
            return {}
    
    async def _fetch_quotes_chunked(self, symbol_mapping: Dict) -> List[Dict]:
        """
        Fetch quotes from Fyers API in chunks (avoid rate limiting).
        Uses chunked batch API calls.
        
        Args:
            symbol_mapping: Dict of underlying -> symbol mapping
            
        Returns:
            List of scanner records with spot + futures prices
        """
        scanner_data = []
        
        try:
            token = self.fyers_auth.get_access_token()
            if not token:
                raise Exception("Not authenticated")
            
            # Prepare all symbols to fetch
            all_symbols = []
            for underlying, symbols in symbol_mapping.items():
                all_symbols.append(symbols['spot'])
                if symbols['future1']:
                    all_symbols.append(symbols['future1'])
                if symbols['future2']:
                    all_symbols.append(symbols['future2'])
                if symbols['future3']:
                    all_symbols.append(symbols['future3'])
            
            logger.info(f"Fetching {len(all_symbols)} symbols from Fyers API")
            
            # Chunk size (Fyers typically allows 50-100 per request)
            chunk_size = 50
            chunks = [all_symbols[i:i+chunk_size] for i in range(0, len(all_symbols), chunk_size)]
            
            # Store quote data
            quotes_data = {}
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                for i, chunk in enumerate(chunks):
                    try:
                        logger.info(f"Fetching chunk {i+1}/{len(chunks)}...")
                        
                        # Build quote request
                        payload = {
                            "symbols": chunk
                        }
                        
                        headers = {
                            "Authorization": f"Bearer {token}"
                        }
                        
                        response = await client.get(
                            self.FYERS_QUOTE_API,
                            json=payload,
                            headers=headers
                        )
                        
                        if response.status_code == 200:
                            quotes = response.json().get('data', {})
                            quotes_data.update(quotes)
                            logger.info(f"Chunk {i+1} fetched: {len(quotes)} quotes")
                        else:
                            logger.warning(f"Chunk {i+1} failed: {response.status_code}")
                        
                        # Rate limit: small delay between chunks
                        await asyncio.sleep(0.5)
                    
                    except Exception as e:
                        logger.warning(f"Error fetching chunk {i+1}: {e}")
                        continue
            
            # Build final scanner data
            for underlying, symbols in symbol_mapping.items():
                try:
                    spot_quote = quotes_data.get(symbols['spot'], {})
                    future1_quote = quotes_data.get(symbols['future1'], {}) if symbols['future1'] else {}
                    future2_quote = quotes_data.get(symbols['future2'], {}) if symbols['future2'] else {}
                    future3_quote = quotes_data.get(symbols['future3'], {}) if symbols['future3'] else {}
                    
                    # Extract LTP (Last Traded Price)
                    spot_price = spot_quote.get('ltp', spot_quote.get('lastPrice', 0))
                    future1_price = future1_quote.get('ltp', future1_quote.get('lastPrice', 0))
                    future2_price = future2_quote.get('ltp', future2_quote.get('lastPrice', 0))
                    future3_price = future3_quote.get('ltp', future3_quote.get('lastPrice', 0))
                    
                    record = {
                        'nse_symbol': underlying,
                        'spot': spot_price if spot_price else '-',
                        'future1': future1_price if future1_price else '-',
                        'future2': future2_price if future2_price else '-',
                        'future3': future3_price if future3_price else '-',
                        'timestamp': datetime.now().isoformat()
                    }
                    
                    scanner_data.append(record)
                
                except Exception as e:
                    logger.warning(f"Error processing quotes for {underlying}: {e}")
                    continue
            
            # Sort by symbol
            scanner_data.sort(key=lambda x: x['nse_symbol'])
            
            logger.info(f"Scanner data prepared: {len(scanner_data)} records")
            return scanner_data
        
        except Exception as e:
            logger.error(f"Quote fetching error: {e}", exc_info=True)
            raise
    
    def get_cached_data(self) -> List[Dict]:
        """
        Get currently cached scanner data.
        
        Returns:
            List of cached FUTSTK records
        """
        return self.cached_data
    
    def get_cache_timestamp(self) -> Optional[str]:
        """
        Get timestamp of last cache update.
        
        Returns:
            ISO format timestamp string
        """
        if self.cache_timestamp:
            return self.cache_timestamp.isoformat()
        return None
    
    def export_to_excel(self, data: List[Dict]) -> io.BytesIO:
        """
        Export scanner data to Excel (.xlsx) format.
        
        Args:
            data: List of scanner records
            
        Returns:
            BytesIO buffer with Excel file
        """
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "F&O Scanner"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = ["NSE Symbol", "Spot", "Future1", "Future2", "Future3"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Write data rows
            for row_idx, record in enumerate(data, start=2):
                ws.cell(row=row_idx, column=1).value = record.get('nse_symbol', '')
                ws.cell(row=row_idx, column=2).value = record.get('spot', '-')
                ws.cell(row=row_idx, column=3).value = record.get('future1', '-')
                ws.cell(row=row_idx, column=4).value = record.get('future2', '-')
                ws.cell(row=row_idx, column=5).value = record.get('future3', '-')
                
                # Apply borders and alignment
                for col_idx in range(1, 6):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    if col_idx > 1:
                        cell.alignment = Alignment(horizontal="right")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 18
            for col in ['B', 'C', 'D', 'E']:
                ws.column_dimensions[col].width = 15
            
            # Save to BytesIO
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            logger.info(f"Excel file created: {len(data)} records")
            return buffer
        
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            raise
